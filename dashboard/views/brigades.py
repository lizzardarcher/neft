from datetime import date, timedelta
import calendar
import json

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.core.paginator import Paginator
from django.db.models import Q, Count, When, Sum, OuterRef, Subquery, ExpressionWrapper, F, Max
from django.db.models import Value, IntegerField, Case
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse, reverse_lazy
from django.utils.timezone import datetime

from dashboard.forms import BrigadeForm, BrigadeActivityForm, WorkObjectForm, BrigadeRequirementForm
from dashboard.mixins import StaffOnlyMixin
from dashboard.models import Brigade, Equipment, Manufacturer, WorkerActivity, BrigadeActivity, WorkObject, UserProfile, \
    Category, BrigadeEquipmentRequirement
from dashboard.utils.utils import get_days_in_month
from openpyxl import Workbook
from openpyxl.styles import Font

NEGATIVE_BRIGADE_ACTIVITY_TYPES = {'Переезд', 'Простой', 'Авария', 'Движка', '-'}
BRIGADE_MARKED_VALUES = {'own', 'external'}
ZBS_ACTIVITY_PREFIX = '(ЗБС)'
VNS_ACTIVITY_PREFIX = '(ВНС)'

MONTH_CHOICES_RU = tuple(
    zip(
        range(1, 13),
        [
            'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
        ],
    )
)

def _clip_period_to_today(start_date, end_date):
    today = date.today()
    effective_end = min(end_date, today)
    if effective_end < start_date:
        return start_date, start_date - timedelta(days=1), 0
    return start_date, effective_end, (effective_end - start_date).days + 1

def _resolve_period(request):
    mode = request.GET.get('mode', 'month')
    today = date.today()

    if mode == 'range':
        start_raw = request.GET.get('start_date')
        end_raw = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            mode = 'month'
        else:
            if end_date < start_date:
                start_date, end_date = end_date, start_date
            return mode, start_date, end_date

    month = request.GET.get('month', today.month)
    year = request.GET.get('year', today.year)
    try:
        month = int(month)
        year = int(year)
    except (TypeError, ValueError):
        month = today.month
        year = today.year

    if not 1 <= month <= 12:
        month = today.month
    if not 1900 <= year <= today.year + 10:
        year = today.year

    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])
    return 'month', start_date, end_date



def _get_marked_brigades_queryset(base_queryset=None):
    queryset = base_queryset if base_queryset is not None else Brigade.objects.all()
    if hasattr(Brigade, 'affiliation'):
        return queryset.filter(affiliation__in=BRIGADE_MARKED_VALUES)
    return queryset


def _empty_brigade_load_stats(brigade, period_start, period_end):
    return {
        'brigade': brigade,
        'period_start': period_start,
        'period_end': period_end,
        'days_in_period': 0,
        'days_total': 0,
        'days_denominator': 0,
        'days_with_any_activity': 0,
        'unmarked_days': 0,
        'positive_days': 0,
        'zbs_days': 0,
        'vns_days': 0,
        'pereezd_days': 0,
        'dvizhka_days': 0,
        'prostoy_days': 0,
        'avarya_days': 0,
        'total_accounted_days': 0,
        'load_percent': 0,
        'load_ratio': 0.0,
        'by_work_type': [],
    }


def _build_brigade_load_stats(brigade, start_date, end_date):
    """
    Загрузка по бригаде: знаменатель D — число фактически размеченных дней в периоде
    (по одной последней записи на день, выбираем max(id) на дату).
    Числитель — дни с «рабочим» типом (все не из NEGATIVE_BRIGADE_ACTIVITY_TYPES).
    Дополнительно: сутки ЗБС / ВНС / переезд / движка / простой / авария.
    """
    period_start, period_end, days_in_period = _clip_period_to_today(start_date, end_date)

    if period_end < period_start:
        return _empty_brigade_load_stats(brigade, period_start, period_end)

    latest_ids_per_day = BrigadeActivity.objects.filter(
        brigade=brigade,
        date__range=(period_start, period_end),
    ).values('date').annotate(last_id=Max('id')).values('last_id')

    effective_qs = BrigadeActivity.objects.filter(id__in=Subquery(latest_ids_per_day)).order_by('date')
    effective_list = list(effective_qs)

    if not effective_list:
        stats = _empty_brigade_load_stats(brigade, period_start, period_end)
        stats['period_start'] = period_start
        stats['period_end'] = period_end
        stats['days_in_period'] = days_in_period
        stats['unmarked_days'] = days_in_period
        return stats

    D = len(effective_list)
    unmarked_days = max(days_in_period - D, 0)
    zbs = vns = pereezd = prostoy = dvizhka = avarya = 0
    positive_days = 0

    for act in effective_list:
        wt = act.work_type
        if wt.startswith(ZBS_ACTIVITY_PREFIX):
            zbs += 1
        elif wt.startswith(VNS_ACTIVITY_PREFIX):
            vns += 1
        elif wt == 'Переезд':
            pereezd += 1
        elif wt == 'Простой':
            prostoy += 1
        elif wt == 'Движка':
            dvizhka += 1
        elif wt == 'Авария':
            avarya += 1

        if wt not in NEGATIVE_BRIGADE_ACTIVITY_TYPES:
            positive_days += 1

    load_ratio = (positive_days / D) if D else 0.0
    load_percent = round(load_ratio * 100, 2)

    by_work_type = list(
        effective_qs.values('work_type')
        .annotate(total=Count('id'))
        .order_by('-total', 'work_type')
    )
    for row in by_work_type:
        row['is_positive'] = row['work_type'] not in NEGATIVE_BRIGADE_ACTIVITY_TYPES

    return {
        'brigade': brigade,
        'period_start': period_start,
        'period_end': period_end,
        'days_in_period': days_in_period,
        'days_total': D,
        'days_denominator': D,
        'days_with_any_activity': len(effective_list),
        'unmarked_days': unmarked_days,
        'positive_days': positive_days,
        'zbs_days': zbs,
        'vns_days': vns,
        'pereezd_days': pereezd,
        'dvizhka_days': dvizhka,
        'prostoy_days': prostoy,
        'avarya_days': avarya,
        'total_accounted_days': D,
        'load_percent': load_percent,
        'load_ratio': load_ratio,
        'by_work_type': by_work_type,
    }


def _affiliation_day_worktypes_in_period(affiliation, period_start, period_end):
    bids = list(Brigade.objects.filter(affiliation=affiliation).values_list('id', flat=True))
    if not bids:
        return {}, bids

    per_day_brigade = (
        BrigadeActivity.objects.filter(
            brigade_id__in=bids,
            date__range=(period_start, period_end),
        )
        .values('brigade_id', 'date')
        .annotate(last_id=Max('id'))
    )
    id_list = [row['last_id'] for row in per_day_brigade]
    cache = {}
    for act in BrigadeActivity.objects.filter(id__in=id_list):
        cache[(act.brigade_id, act.date)] = act.work_type
    return cache, bids


def _daily_affiliation_load_percent_series(affiliation, period_start, period_end):
    """
    По каждому календарному дню периода: доля бригад с положительной последней активностью дня, в процентах 0–100.
    Знаменатель — число бригад данной аффилиации; при 0 бригадах для дня — None.
    """
    cache, bids = _affiliation_day_worktypes_in_period(affiliation, period_start, period_end)
    n = len(bids)
    if n == 0:
        return [None] * ((period_end - period_start).days + 1 if period_end >= period_start else 0)

    out = []
    d = period_start
    while d <= period_end:
        good = 0
        for bid in bids:
            wt = cache.get((bid, d))
            if wt and wt not in NEGATIVE_BRIGADE_ACTIVITY_TYPES:
                good += 1
        out.append(round((good / n) * 100, 2))
        d += timedelta(days=1)
    return out


def _last_activity_date_for_affiliation(affiliation, period_start, period_end):
    return (
        BrigadeActivity.objects.filter(
            brigade__affiliation=affiliation,
            date__range=(period_start, period_end),
        )
        .order_by('-date')
        .values_list('date', flat=True)
        .first()
    )


def _mean_positive(values):
    usable = [v for v in values if v is not None]
    if not usable:
        return None
    return round(sum(usable) / len(usable), 2)


def _load_percent_css_class(p):
    if p is None:
        return 'text-muted'
    if p < 65:
        return 'text-danger'
    if p < 80:
        return 'text-warning'
    return 'text-success'


def _excel_font_for_load_percent(p):
    if p is None:
        return Font(color='666666')
    if p < 65:
        return Font(color='CC3333')
    if p < 80:
        return Font(color='C47A00')
    return Font(color='198754')


def _load_percent_bar_color(p):
    if p < 65:
        return '#dc3545'
    if p < 80:
        return '#ffc107'
    return '#198754'


def _affiliation_chart_label_color(affiliation):
    if affiliation == 'own':
        return '#198754'
    if affiliation == 'external':
        return '#fd7e14'
    return '#6c757d'


def build_org_daily_affiliation_display(start_date, end_date):
    """
    Данные для таблицы «% загрузки собственные / субподряд» по дням выбранного периода (учёт обрезки по сегодня).
    """
    period_start, period_end, span = _clip_period_to_today(start_date, end_date)
    if span == 0 or period_end < period_start:
        return {
            'org_daily_headers': [],
            'org_daily_own_cells': [],
            'org_daily_external_cells': [],
            'org_daily_own_avg': None,
            'org_daily_external_avg': None,
            'org_daily_clipped_label': f'{period_start} — {period_end}',
        }

    own_last = _last_activity_date_for_affiliation('own', period_start, period_end)
    ext_last = _last_activity_date_for_affiliation('external', period_start, period_end)
    display_end = max([d for d in [own_last, ext_last] if d is not None], default=None)
    if display_end is None:
        return {
            'org_daily_headers': [],
            'org_daily_own_cells': [],
            'org_daily_external_cells': [],
            'org_daily_own_avg': None,
            'org_daily_external_avg': None,
            'org_daily_clipped_label': f'{period_start} — {period_end}',
        }

    own_raw = (
        _daily_affiliation_load_percent_series('own', period_start, own_last)
        if own_last is not None else []
    )
    ext_raw = (
        _daily_affiliation_load_percent_series('external', period_start, ext_last)
        if ext_last is not None else []
    )
    own_avg = _mean_positive(own_raw)
    ext_avg = _mean_positive(ext_raw)

    headers = []
    own_cells = []
    ext_cells = []
    d = period_start
    i = 0
    while d <= display_end:
        headers.append(d.strftime('%d.%m'))
        op = own_raw[i] if i < len(own_raw) else None
        ep = ext_raw[i] if i < len(ext_raw) else None
        own_cells.append(
            {'value': op, 'css_class': _load_percent_css_class(op), 'suffix': '' if op is None else '%'}
        )
        ext_cells.append(
            {'value': ep, 'css_class': _load_percent_css_class(ep), 'suffix': '' if ep is None else '%'}
        )
        d += timedelta(days=1)
        i += 1

    if headers:
        own_cells.append(
            {
                'value': own_avg,
                'css_class': _load_percent_css_class(own_avg) if own_avg is not None else 'text-muted',
                'suffix': '%' if own_avg is not None else '',
            }
        )
        ext_cells.append(
            {
                'value': ext_avg,
                'css_class': _load_percent_css_class(ext_avg) if ext_avg is not None else 'text-muted',
                'suffix': '%' if ext_avg is not None else '',
            }
        )

    header_labels = headers + (['Итог'] if headers else [])

    return {
        'org_daily_headers': header_labels,
        'org_daily_own_cells': own_cells,
        'org_daily_external_cells': ext_cells,
        'org_daily_own_avg': own_avg,
        'org_daily_external_avg': ext_avg,
        'org_daily_clipped_label': f'{period_start} — {display_end}',
    }

# class BrigadeListView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, ListView):
#     model = Brigade
#     context_object_name = 'brigades'
#     template_name = 'dashboard/brigades/brigade_list.html'
#     ordering = 'name'
#
#     def get_queryset(self):
#         """Поиск по имени и описанию бригады"""
#         queryset = super().get_queryset()
#         search_request = self.request.GET.get("search")
#         if search_request:
#             brigade_by_name = Brigade.objects.filter(name__icontains=search_request)
#             brigade_by_description = Brigade.objects.filter(description__icontains=search_request)
#             queryset = brigade_by_description | brigade_by_name
#         return queryset
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['page_size'] = self.request.GET.get('page_size', self.paginate_by)
#         return context


class BrigadeListView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, ListView):
    model = Brigade
    context_object_name = 'brigades'
    template_name = 'dashboard/brigades/brigade_list.html'
    ordering = 'name'

    def get_queryset(self):
        """Поиск по имени и описанию бригады"""
        queryset = super().get_queryset()

        required_qty_subquery = BrigadeEquipmentRequirement.objects.filter(
            brigade=OuterRef('pk'),
            category=OuterRef('category')
        ).values('brigade').annotate(
            total_sum_qty=Sum('quantity')
        ).values('total_sum_qty')[:1]

        queryset = queryset.annotate(
        )

        search_request = self.request.GET.get("search")
        if search_request:
            queryset = queryset.filter(
                Q(name__icontains=search_request) |
                Q(description__icontains=search_request)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_size'] = self.request.GET.get('page_size', self.paginate_by)

        total_shortage_all_brigades = 0
        total_surplus_all_brigades = 0

        brigades_with_equipment = []
        for brigade in context['brigades']:  # context['brigades'] уже пагинирован
            equipment_status = self._get_equipment_status(brigade)
            brigade.equipment_info = equipment_status
            brigades_with_equipment.append(brigade)

            total_shortage_all_brigades += equipment_status['shortage_total']
            total_surplus_all_brigades += equipment_status['surplus_total']

        context['brigades'] = brigades_with_equipment
        context['total_shortage_all_brigades'] = total_shortage_all_brigades
        context['total_surplus_all_brigades'] = total_surplus_all_brigades

        return context

    def _get_equipment_status(self, brigade):
        """
        Подсчитывает недостающее и лишнее оборудование для бригады.
        Оптимизированная версия с использованием агрегаций Django ORM.
        Возвращает словарь с ключами: shortage_total, surplus_total
        """
        required_subquery = BrigadeEquipmentRequirement.objects.filter(
            brigade=brigade,
            category=OuterRef('pk')
        ).values('category').annotate(
            total_required_for_category=Sum('quantity')
        ).values('total_required_for_category')[:1]

        category_stats = Category.objects.annotate(
            required_qty=Coalesce(Subquery(required_subquery, output_field=IntegerField()), Value(0)),
            actual_qty=Count(
                'equipment',
                filter=Q(equipment__brigade=brigade),  # Фильтруем оборудование, относящееся к данной бригаде
                distinct=True
            )
        ).annotate(
            difference=F('actual_qty') - F('required_qty'),
            shortage_cat=Case(
                When(difference__lt=0, then=ExpressionWrapper(F('difference') * -1, output_field=IntegerField())),
                default=Value(0),
                output_field=IntegerField()
            ),
            surplus_cat=Case(
                When(difference__gt=0, then=F('difference')),
                default=Value(0),
                output_field=IntegerField()
            )
        ).values('shortage_cat', 'surplus_cat')

        total_summary = category_stats.aggregate(
            shortage_total=Sum('shortage_cat', default=0),
            surplus_total=Sum('surplus_cat', default=0)
        )

        return {
            'shortage_total': total_summary['shortage_total'],
            'surplus_total': total_summary['surplus_total']
        }

class BrigadeDetailView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, DetailView):
    model = Brigade
    context_object_name = 'brigade'
    template_name = 'dashboard/brigades/brigade_detail.html'

    def get_context_data(self, **kwargs):
        equipments = Equipment.objects.filter(brigade=self.get_object()).select_related('category', 'brigade').prefetch_related('documents')
        search_request = self.request.GET.get("search")
        category = self.request.GET.get("category")
        sort_by = self.request.GET.get('sort_by', 'id')  # по умолчанию сортируем по id
        order = self.request.GET.get('order', 'asc')  # по умолчанию прямой порядок
        if search_request:
            equipment_by_name = Equipment.objects.filter(brigade=self.get_object(), name__icontains=search_request).select_related('category', 'brigade').prefetch_related('documents')
            equipment_by_serial = Equipment.objects.filter(brigade=self.get_object(), serial__icontains=search_request).select_related('category', 'brigade').prefetch_related('documents')
            equipments = (equipment_by_name | equipment_by_serial)

        if category:
            equipments = equipments.filter(category__name=category)

        if sort_by:
            if order == 'desc':
                sort_by = f"-{sort_by}"
                equipments = equipments.order_by(sort_by)
            else:
                equipments = equipments.order_by(sort_by)

        # Добавляем пагинацию
        paginator = Paginator(equipments, 30)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = super().get_context_data(**kwargs)
        context['equipments'] = page_obj
        context['page_obj'] = page_obj
        context['now_date'] = date.today()
        context['manufacturers'] = Manufacturer.objects.all().order_by('name')

        return context


class BrigadeStaffView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/brigades/brigade_staff.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        brigade = get_object_or_404(Brigade, pk=self.request.path.split('/')[-2])
        context['brigade'] = brigade
        context['month'] = self.request.GET.get('month', datetime.now().strftime('%m'))
        context['year'] = self.request.GET.get('year', datetime.now().strftime('%Y'))
        context['days'] = get_days_in_month(int(context['month']), int(context['year']))
        context['date_m_y'] = datetime.now().strftime('%m-%Y')
        prev_month = int(context['month']) - 1
        next_month = int(context['month']) + 1
        prev_year = int(context['year'])
        next_year = int(context['year'])
        if prev_month < 1:
            prev_month = 12
            prev_year = int(context['year']) - 1
        if str(prev_month).__len__() == 1:
            prev_month = '0' + str(prev_month)
        if next_month > 12:
            next_month = 1
            next_year = int(context['year']) + 1
        if str(next_month).__len__() == 1:
            next_month = '0' + str(next_month)
        context['prev_month'] = str(prev_month)
        context['next_month'] = str(next_month)
        context['prev_year'] = prev_year
        context['next_year'] = next_year

        # 1. Пользователи, которые сейчас числятся в бригаде
        current_brigade_users = User.objects.filter(
            profile__brigade=brigade,
            is_staff=True,
            workeractivity__date__year=context['year'],
            workeractivity__date__month=context['month'],
            workeractivity__brigade=brigade
        ).annotate(
            has_wa=Case(
                When(workeractivity__date__year=context['year'], workeractivity__date__month=context['month'],
                     workeractivity__brigade=brigade, then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            ),
            total_wa=Count('workeractivity',
                           filter=Q(workeractivity__date__year=context['year'], workeractivity__date__month=context['month'],
                                    workeractivity__brigade=brigade))
        )

        # 2. Пользователи, которые когда-либо работали в бригаде (через WorkerActivity)
        past_brigade_users = User.objects.filter(
            is_staff=True,
            workeractivity__date__year=context['year'],
            workeractivity__date__month=context['month'],
            workeractivity__brigade=brigade
        ).annotate(
            has_wa=Case(
                When(workeractivity__date__year=context['year'], workeractivity__date__month=context['month'],
                     workeractivity__brigade=brigade, then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            ),
            total_wa=Count('workeractivity',
                           filter=Q(workeractivity__date__year=context['year'], workeractivity__date__month=context['month'],
                                    workeractivity__brigade=brigade))
        )

        # Объединяем два queryset-а и убираем дубликаты
        all_brigade_users = current_brigade_users.union(past_brigade_users).order_by('-has_wa',
                                                                                     'username')  # Убираем distinct(), так как union уже делает это

        context['users'] = all_brigade_users
        context['brigade_users'] = current_brigade_users
        employee_data = [
            {
                'user': user,
                'total_wa': WorkerActivity.objects.filter(user=user, date__month=context['month'],
                                                          date__year=context['year'], brigade=brigade).count(),
                'wa': [
                    {'day': day, 'wa': WorkerActivity.objects.filter(user=user, date__month=context['month'],
                                                                     date__year=context['year'], date__day=day,
                                                                     brigade=brigade).last()}
                    for day in context['days']
                ],
            } for user in all_brigade_users
        ]
        context['employee_data'] = employee_data


        return context


class BrigadeWorkView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/brigades/brigade_work.html'

    def form_valid(self, form):
        form = BrigadeActivityForm(brigade_id=int(self.request.path.split('/')[-5]), data=self.request.POST)
        if form.is_valid():
            form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['form'] = BrigadeActivityForm()
        context['work_object_form'] = WorkObjectForm()
        brigade_id = int(self.request.path.split('/')[-5])
        year = int(self.kwargs.get('year', datetime.now().year))
        month = int(self.kwargs.get('month', datetime.now().month))

        brigade = get_object_or_404(Brigade, pk=brigade_id)
        context['brigade'] = brigade

        activities = BrigadeActivity.objects.filter(
            brigade=brigade,
            date__year=year,
            date__month=month
        )
        work_objects = []
        for i in activities:
            if i.work_object:
                work_objects.append(i.work_object)
        context['work_objects'] = work_objects
        # context['work_objects'] = WorkObject.objects.all().order_by('name')

        cal = calendar.monthcalendar(year, month)  # Returns a list of lists
        context['calendar'] = cal

        context['year'] = year
        context['month'] = month if month > 9 else '0' + str(month)
        context['month_name'] = calendar.month_name[month]

        brigade_data = [
            {
                'brigade': brigade,
                'total_ba': BrigadeActivity.objects.filter(brigade=brigade, date__month=month, date__year=year).count(),
                'ba': [
                    {'day': day,
                     'ba': BrigadeActivity.objects.filter(brigade=brigade, date__month=month, date__year=year,
                                                          date__day=day).last()}
                    for day in get_days_in_month(month, year)
                ],
            }
        ]
        context['brigade_data'] = brigade_data
        context['days'] = get_days_in_month(month, year)

        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1

        context['prev_month_url'] = f'/dashboard/brigade/{brigade_id}/work/{prev_month}/{prev_year}'
        context['next_month_url'] = f'/dashboard/brigade/{brigade_id}/work/{next_month}/{next_year}'

        activities_by_day = {}
        for activity in activities:
            day = activity.date.day
            if day not in activities_by_day:
                activities_by_day[day] = []
            activities_by_day[day].append(activity)

        context['activities_by_day'] = activities_by_day

        return context


class BrigadeTableTotalView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/brigades/brigade_work_total.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        month = int(self.request.GET.get('month', datetime.now().month))
        year = int(self.request.GET.get('year', datetime.now().year))
        brigades = Brigade.objects.all().order_by('name')

        if not 1 <= month <= 12:
            month = datetime.now().month
        if not 1900 <= year <= datetime.now().year + 10:
            year = datetime.now().year

        brigade_data = []
        for brigade in brigades:
            day_strings = get_days_in_month(month, year)
            ba_cells = []
            for day_str in day_strings:
                day_int = int(day_str)
                ba_cells.append(
                    {
                        'day': day_str,
                        'ba': BrigadeActivity.objects.filter(
                            brigade=brigade,
                            date__month=month,
                            date__year=year,
                            date__day=day_int,
                        ).last(),
                    }
                )
            brigade_data.append(
                {
                    'brigade': brigade,
                    'total_ba': BrigadeActivity.objects.filter(
                        brigade=brigade, date__month=month, date__year=year
                    ).count(),
                    'ba': ba_cells,
                }
            )

        context['brigade_data'] = brigade_data

        context['form'] = BrigadeActivityForm()
        context['work_object_form'] = WorkObjectForm()
        work_objects = []
        for i in BrigadeActivity.objects.filter(date__month=month, date__year=year):
            if i.work_object:
                work_objects.append(i.work_object)
        work_objects = set(work_objects)

        context['work_objects'] = work_objects
        # context['work_objects'] = WorkObject.objects.all().order_by('name')
        context['month'] = month
        context['year'] = year
        context['current_month'] = datetime.now().month
        context['current_year'] = datetime.now().year
        context['days'] = get_days_in_month(month, year)
        context['brigades'] = brigades
        context['days_in_month'] = calendar.monthrange(year, month)[1]
        context['previous_month_url'] = self.get_month_url(month, year, -1)
        context['next_month_url'] = self.get_month_url(month, year, 1)

        return context

    def get_month_url(self, month, year, offset):
        new_month = month + offset
        new_year = year

        if new_month > 12:
            new_month = 1
            new_year += 1
        elif new_month < 1:
            new_month = 12
            new_year -= 1

        return reverse('brigade_table_total') + f'?month={new_month}&year={new_year}'


class BrigadeIndexView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/brigades/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['month'] = datetime.now().strftime('%m')
        context['year'] = datetime.now().strftime('%Y')
        context['brigade'] = get_object_or_404(Brigade, pk=self.request.path.split('/')[-1])
        context['staff_count'] = User.objects.filter(profile__brigade=context['brigade']).count()
        context['equipment_count'] = Equipment.objects.filter(brigade=context['brigade']).count()
        return context


class BrigadeLoadAnalysisView(LoginRequiredMixin, StaffOnlyMixin, TemplateView):
    template_name = 'dashboard/brigades/brigade_load_analysis.html'

    def dispatch(self, request, *args, **kwargs):
        self.brigade = get_object_or_404(Brigade, pk=kwargs['pk'])
        if hasattr(self.brigade, 'affiliation') and self.brigade.affiliation not in BRIGADE_MARKED_VALUES:
            messages.warning(request, 'Проставьте тип бригады: своя или чужая, чтобы открыть анализ.')
            return redirect(reverse('brigade_table_total') + f'?month={date.today().month}&year={date.today().year}')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mode, start_date, end_date = _resolve_period(self.request)
        stats = _build_brigade_load_stats(self.brigade, start_date, end_date)

        context.update({
            'brigade': self.brigade,
            'stats': stats,
            'mode': mode,
            'start_date_input': start_date.strftime('%Y-%m-%d'),
            'end_date_input': end_date.strftime('%Y-%m-%d'),
            'month_input': start_date.month,
            'year_input': start_date.year,
            'month_choices': MONTH_CHOICES_RU,
        })
        return context


class OrganizationLoadAnalysisView(LoginRequiredMixin, StaffOnlyMixin, TemplateView):
    template_name = 'dashboard/brigades/organization_load_analysis.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mode, start_date, end_date = _resolve_period(self.request)
        _, _, days_in_period = _clip_period_to_today(start_date, end_date)
        brigades = _get_marked_brigades_queryset(Brigade.objects.all().order_by('name'))
        rows = [_build_brigade_load_stats(brigade, start_date, end_date) for brigade in brigades]

        total_positive_days = sum(row['positive_days'] for row in rows)
        org_capacity_days = sum(row['days_total'] for row in rows)
        brigades_in_report = len(rows)
        org_percent = round((total_positive_days / org_capacity_days) * 100, 2) if org_capacity_days else 0
        own_rows = [row for row in rows if getattr(row['brigade'], 'affiliation', '') == 'own']
        external_rows = [row for row in rows if getattr(row['brigade'], 'affiliation', '') == 'external']
        own_positive_days = sum(row['positive_days'] for row in own_rows)
        own_count = len(own_rows)
        own_capacity = sum(row['days_total'] for row in own_rows)
        own_percent = round((own_positive_days / own_capacity) * 100, 2) if own_capacity else 0
        external_positive_days = sum(row['positive_days'] for row in external_rows)
        external_count = len(external_rows)
        external_capacity = sum(row['days_total'] for row in external_rows)
        external_percent = round((external_positive_days / external_capacity) * 100, 2) if external_capacity else 0
        org_table_totals = None
        if rows:
            org_table_totals = {
                'days_total': sum(r['days_total'] for r in rows),
                'unmarked_days': sum(r.get('unmarked_days', 0) for r in rows),
                'zbs_days': sum(r.get('zbs_days', 0) for r in rows),
                'vns_days': sum(r.get('vns_days', 0) for r in rows),
                'positive_days': sum(r['positive_days'] for r in rows),
                'pereezd_days': sum(r.get('pereezd_days', 0) for r in rows),
                'dvizhka_days': sum(r.get('dvizhka_days', 0) for r in rows),
                'prostoy_days': sum(r.get('prostoy_days', 0) for r in rows),
            }
        context.update({
            'rows': rows,
            'org_table_totals': org_table_totals,
            'mode': mode,
            'start_date_input': start_date.strftime('%Y-%m-%d'),
            'end_date_input': end_date.strftime('%Y-%m-%d'),
            'month_input': start_date.month,
            'year_input': start_date.year,
            'month_choices': MONTH_CHOICES_RU,
            'organization_positive_days': total_positive_days,
            'organization_calendar_days': days_in_period,
            'organization_brigades_count': brigades_in_report,
            'organization_capacity_days': org_capacity_days,
            'organization_load_percent': org_percent,
            'own_positive_days': own_positive_days,
            'own_brigades_count': own_count,
            'own_capacity_days': own_capacity,
            'own_load_percent': own_percent,
            'external_positive_days': external_positive_days,
            'external_brigades_count': external_count,
            'external_capacity_days': external_capacity,
            'external_load_percent': external_percent,
            'chart_labels_json': json.dumps([row['brigade'].name for row in rows], ensure_ascii=False),
            'chart_data_json': json.dumps([row['load_percent'] for row in rows]),
            'chart_bar_colors_json': json.dumps([
                _load_percent_bar_color(row['load_percent']) for row in rows
            ]),
            'chart_label_colors_json': json.dumps([
                _affiliation_chart_label_color(getattr(row['brigade'], 'affiliation', ''))
                for row in rows
            ]),
        })
        context.update(build_org_daily_affiliation_display(start_date, end_date))
        return context


def export_brigade_load_excel(request, pk):
    brigade = get_object_or_404(Brigade, pk=pk)
    mode, start_date, end_date = _resolve_period(request)
    stats = _build_brigade_load_stats(brigade, start_date, end_date)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Загрузка {brigade.name[:20]}"
    sheet.append(['Бригада', brigade.name])
    sheet.append(['Период', f"{stats['period_start']} - {stats['period_end']}"])
    sheet.append(
        [
            'ЗБС',
            'ВНС',
            'ВСЕГО',
            'Переезд',
            'Движка',
            'Простой',
            'Авария',
            'Итого суток',
            'Неразмечено',
            '%',
        ]
    )
    sheet.append(
        [
            stats.get('zbs_days', 0),
            stats.get('vns_days', 0),
            stats['positive_days'],
            stats.get('pereezd_days', 0),
            stats.get('dvizhka_days', 0),
            stats.get('prostoy_days', 0),
            stats.get('avarya_days', 0),
            stats.get('total_accounted_days', stats['days_total']),
            stats.get('unmarked_days', 0),
            stats['load_percent'],
        ]
    )
    sheet.append(['D', stats['days_total']])
    sheet.append([])
    sheet.append(['Тип активности', 'Количество дней', 'Положительная'])
    for row in stats['by_work_type']:
        sheet.append([row['work_type'], row['total'], 'Да' if row['is_positive'] else 'Нет'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=brigade_load_{brigade.id}.xlsx'
    workbook.save(response)
    return response


def export_organization_load_excel(request):
    mode, start_date, end_date = _resolve_period(request)
    brigades = _get_marked_brigades_queryset(Brigade.objects.all().order_by('name'))
    rows = [_build_brigade_load_stats(brigade, start_date, end_date) for brigade in brigades]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Организация"
    sheet.append(['Период', f"{start_date} - {end_date}"])
    sheet.append(
        [
            'Бригада',
            'D',
            'Неразмечено',
            'ЗБС',
            'ВНС',
            'ВСЕГО',
            'Переезд',
            'Движка',
            'Простой',
            '%',
        ]
    )
    for row in rows:
        b = row['brigade']
        aff = getattr(b, 'affiliation', '') or ''
        prefix = '[С] ' if aff == 'own' else '[Ч] ' if aff == 'external' else ''
        sheet.append(
            [
                f"{prefix}{b.name}",
                row['days_total'],
                row.get('unmarked_days', 0),
                row.get('zbs_days', 0),
                row.get('vns_days', 0),
                row['positive_days'],
                row.get('pereezd_days', 0),
                row.get('dvizhka_days', 0),
                row.get('prostoy_days', 0),
                row['load_percent'],
            ]
        )

    if rows:
        sheet.append(
            [
                'Итого',
                sum(r['days_total'] for r in rows),
                sum(r.get('unmarked_days', 0) for r in rows),
                sum(r.get('zbs_days', 0) for r in rows),
                sum(r.get('vns_days', 0) for r in rows),
                sum(r['positive_days'] for r in rows),
                sum(r.get('pereezd_days', 0) for r in rows),
                sum(r.get('dvizhka_days', 0) for r in rows),
                sum(r.get('prostoy_days', 0) for r in rows),
                '',
            ]
        )

    total_positive_days = sum(row['positive_days'] for row in rows)
    org_capacity = sum(row['days_total'] for row in rows)
    total_percent = round((total_positive_days / org_capacity) * 100, 2) if org_capacity else 0
    own_rows = [row for row in rows if getattr(row['brigade'], 'affiliation', '') == 'own']
    external_rows = [row for row in rows if getattr(row['brigade'], 'affiliation', '') == 'external']
    own_positive_days = sum(row['positive_days'] for row in own_rows)
    own_capacity = sum(row['days_total'] for row in own_rows)
    own_percent = round((own_positive_days / own_capacity) * 100, 2) if own_capacity else 0
    external_positive_days = sum(row['positive_days'] for row in external_rows)
    external_capacity = sum(row['days_total'] for row in external_rows)
    external_percent = round((external_positive_days / external_capacity) * 100, 2) if external_capacity else 0
    sheet.append([])
    sheet.append(['Свод', 'Рабочих дней', 'Календарных дней', 'Загрузка %'])
    sheet.append(['Организация', total_positive_days, org_capacity, total_percent])
    sheet.append(['Свои бригады', own_positive_days, own_capacity, own_percent])
    sheet.append(['Чужие бригады', external_positive_days, external_capacity, external_percent])

    daily = build_org_daily_affiliation_display(start_date, end_date)
    if daily['org_daily_headers']:
        sheet.append([])
        sheet.append(['Загрузка по дням', daily['org_daily_clipped_label']])
        header_row = ['Показатель'] + daily['org_daily_headers']
        sheet.append(header_row)

        def _write_pct_row(label, cells):
            row_values = [label] + [
                (cell['value'] if cell.get('value') is not None else '')
                for cell in cells
            ]
            sheet.append(row_values)
            out_row = sheet.max_row
            for col_i, val in enumerate(row_values, start=1):
                if col_i == 1:
                    continue
                if val == '' or val is None:
                    continue
                try:
                    fval = float(val)
                except (TypeError, ValueError):
                    continue
                sheet.cell(row=out_row, column=col_i).font = _excel_font_for_load_percent(fval)

        _write_pct_row('% загрузки собственные', daily['org_daily_own_cells'])
        _write_pct_row('% загрузки субподряд', daily['org_daily_external_cells'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=organization_load.xlsx'
    workbook.save(response)
    return response


class BrigadeCreateView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, CreateView):
    model = Brigade
    context_object_name = 'brigade'
    template_name = 'dashboard/brigades/brigade_form_create.html'
    form_class = BrigadeForm
    success_url = '/dashboard/brigades'

    def get_success_message(self, cleaned_data):
        return f"Бригада {cleaned_data['name']} Успешно создана!"


class BrigadeUpdateView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, UpdateView):
    model = Brigade
    context_object_name = 'brigade'
    template_name = 'dashboard/brigades/brigade_form_update.html'
    form_class = BrigadeForm
    success_url = '/dashboard/brigades'

    def get_success_message(self, cleaned_data):
        return f"Бригада {cleaned_data['name']} Успешно обновлена!"


class BrigadeUpdateFromTotalView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, UpdateView):
    model = Brigade
    context_object_name = 'brigade'
    template_name = 'dashboard/brigades/brigade_form_update.html'
    form_class = BrigadeForm

    def get_success_url(self):
        return reverse('brigade_table_total')

    def get_success_message(self, cleaned_data):
        return f"Бригада {cleaned_data['name']} Успешно обновлена!"


class BrigadeUpdateFromWorkView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, UpdateView):
    model = Brigade
    context_object_name = 'brigade'
    template_name = 'dashboard/brigades/brigade_form_update.html'
    form_class = BrigadeForm

    def get_success_url(self):
        return reverse('brigade_work', args=[self.kwargs.get('pk'), self.kwargs.get('month'), self.kwargs.get('year')])

    def get_success_message(self, cleaned_data):
        return f"Бригада {cleaned_data['name']} Успешно обновлена!"


def brigade_delete(request, brigade_id):
    brigade = get_object_or_404(Brigade, id=brigade_id)
    brigade.delete()
    messages.success(request, 'Бригада успешно удалена!')
    return redirect('brigade_list')


def brigade_activity_create(request, brigade_id):
    form = BrigadeActivityForm(request.POST)
    brigade = get_object_or_404(Brigade, id=brigade_id)
    date = request.POST.get('date')
    work_type = request.POST.get('work_type')
    if work_type == '-':
        BrigadeActivity.objects.filter(brigade=brigade, date=date).delete()
        messages.success(request, 'Активность успешно удалена!')
        return redirect(request.META.get('HTTP_REFERER'))
    else:
        try:
            work_object = get_object_or_404(WorkObject, id=request.POST.get('work_object'))
        except:
            work_object = None

        if brigade and work_type and date:
            brigade_activity = BrigadeActivity.objects.filter(
                brigade=brigade,
                date=date, ).first()
            if brigade_activity:
                brigade_activity.brigade = brigade
                brigade_activity.date = date
                brigade_activity.work_object = work_object
                brigade_activity.work_type = work_type

                # Добавляем работников бригады к активности
                for worker in UserProfile.objects.filter(brigade=brigade):
                    brigade_activity.workers.add(worker.user)

                brigade_activity.save()
                messages.success(request, f'Активность успешно обновлена!')
            else:
                brigade_activity = BrigadeActivity.objects.create(
                    brigade=brigade,
                    date=date,
                    work_object=work_object,
                    work_type=work_type,
                )

                # Добавляем работников бригады к активности
                for worker in UserProfile.objects.filter(
                        brigade=brigade):  # Получаем всех пользователей, связанных с бригадой
                    brigade_activity.workers.add(worker.user)

                messages.success(request, f'Активность успешно создана!')
            return redirect(request.META.get('HTTP_REFERER') + f'#{brigade.id}')
        else:
            if form.is_valid():
                activity = form.save(commit=False)
                activity.brigade = brigade

                # Добавляем работников бригады к активности
                for worker in UserProfile.objects.filter(
                        brigade=brigade):  # Получаем всех пользователей, связанных с бригадой
                    activity.workers.add(worker.user)

                activity.save()
                messages.success(request, 'Активность успешно создана!')
            else:
                messages.error(request, 'Произошла ошибка при создании активности!')
            return redirect(request.META.get('HTTP_REFERER'))


def work_object_create(request):
    if request.method == 'POST':
        form = WorkObjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Объект успешно создан!')
            return redirect(request.META.get('HTTP_REFERER'))
        else:
            messages.error(request, 'Произошла ошибка при создании объекта!')
            return redirect(request.META.get('HTTP_REFERER'))





def get_locations(request):
    query = request.GET.get('term', '')  # Получаем текст, введенный пользователем
    locations = WorkObject.objects.values_list('name', flat=True).distinct()  # Получаем уникальные месторождения
    locations = [loc for loc in locations if query.lower() in loc.lower()]  # Фильтрация для поиска

    data = list(locations)  # Преобразуем в список
    return JsonResponse(data, safe=False)


class WorkObjectListView(LoginRequiredMixin, StaffOnlyMixin, ListView):
    model = WorkObject
    context_object_name = 'work_objects'
    template_name = 'dashboard/brigades/work_object_list.html'


class WorkObjectDetailView(LoginRequiredMixin, StaffOnlyMixin, DetailView):
    model = WorkObject
    context_object_name = 'work_object'
    template_name = 'dashboard/brigades/work_object_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activities'] = BrigadeActivity.objects.filter(work_object=self.object).order_by('date')
        return context


class WorkObjectCreateView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, CreateView):
    model = WorkObject
    template_name = 'dashboard/brigades/work_object_edit.html'
    form_class = WorkObjectForm
    success_message = 'Объект успешно создан!'
    success_url = reverse_lazy('work_object_list')

class WorkObjectUpdateView(LoginRequiredMixin, StaffOnlyMixin, SuccessMessageMixin, UpdateView):
    model = WorkObject
    context_object_name = 'work_object'
    template_name = 'dashboard/brigades/work_object_edit.html'
    form_class = WorkObjectForm
    def get_success_url(self):

        if 'brigade_table_total' in self.request.GET.get('from'):
            return f'/dashboard/brigade/brigade_table_total/?month={datetime.now().month}&year={datetime.now().year}'
        else:
            return reverse('work_object_list')

    def get_success_message(self, cleaned_data):
        return f"Объект {cleaned_data['name']} Успешно обновлен!"

def work_object_delete(request, work_object_id):
    work_object = get_object_or_404(WorkObject, id=work_object_id)
    work_object.delete()
    messages.success(request, 'Объект успешно удален!')
    return redirect(request.META.get('HTTP_REFERER'))


def update_brigade_requirements(request, brigade_id):
    brigade = get_object_or_404(Brigade, id=brigade_id)
    if request.method == 'POST':
        form = BrigadeRequirementForm(request.POST, brigade=brigade)
        if form.is_valid():
            form.save()
            return redirect('brigade_requirement_status', brigade_id=brigade.id)
    else:
        form = BrigadeRequirementForm(brigade=brigade)

    return render(request, 'dashboard/brigades/update_equipment_requirements.html', {'form': form, 'brigade': brigade})


# 2. Страница отображения статуса (Расчет нехватки и лишнего)
def brigade_equipment_status(request, brigade_id):
    brigade = get_object_or_404(Brigade, id=brigade_id)
    categories = Category.objects.all().order_by('name')

    report = []

    # Инициализация переменных для общих итогов
    total_required_all = 0
    total_actual_all = 0
    total_shortage_all = 0
    total_surplus_all = 0

    for cat in categories:
        # План (из нашей связующей модели)
        req = BrigadeEquipmentRequirement.objects.filter(brigade=brigade, category=cat).first()
        required_qty = req.quantity if req else 0

        # Факт (сколько оборудования этой категории сейчас числится за бригадой)
        # Учитываем состояние оборудования
        actual_qty = Equipment.objects.filter(brigade=brigade, category=cat, condition='work').count()

        diff = actual_qty - required_qty

        shortage = abs(diff) if diff < 0 else 0
        surplus = diff if diff > 0 else 0

        report.append({
            'category': cat.name,
            'required': required_qty,
            'actual': actual_qty,
            'shortage': shortage,
            'surplus': surplus
        })

        # Обновляем общие итоги
        total_required_all += required_qty
        total_actual_all += actual_qty
        total_shortage_all += shortage
        total_surplus_all += surplus

    return render(request, 'dashboard/brigades/brigade_equipment_status.html', {
        'brigade': brigade,
        'report': report,
        'page_title': f"Статус оснащения бригады «{brigade.name}»",
        # Передаем общие итоги в контекст
        'total_required_all': total_required_all,
        'total_actual_all': total_actual_all,
        'total_shortage_all': total_shortage_all,
        'total_surplus_all': total_surplus_all,
    })


