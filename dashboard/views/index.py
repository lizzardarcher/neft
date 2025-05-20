import calendar
import io
from datetime import datetime
from re import search

import csv
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.views import View
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl import Workbook

from dashboard.forms import VehicleMovementFilterForm
from dashboard.mixins import StaffOnlyMixin
from dashboard.models import Brigade, Category, Equipment, UserActionLog, Transfer, Document, BrigadeActivity, \
    WorkerActivity, WorkObject, VehicleMovement
from dashboard.utils.utils import get_days_in_month


class DashboardView(LoginRequiredMixin, SuccessMessageMixin, TemplateView):
    template_name = 'dashboard/index.html'

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        context['brigades_count'] = Brigade.objects.all().count()
        context['categories_count'] = Category.objects.all().count()
        context['equipment_count'] = Equipment.objects.all().count()
        context['work_object_count'] = WorkObject.objects.all().count()
        context['users_count'] = User.objects.filter(is_staff=True).count()
        context['user_log_count'] = UserActionLog.objects.all().count()
        context['transfers_count'] = Transfer.objects.all().count()
        context['vehicle_movement_count'] = VehicleMovement.objects.all().count()
        context['document_count'] = Document.objects.all().count()
        context['month'] = datetime.now().strftime('%m')
        context['year'] = datetime.now().strftime('%Y')
        return context


class InstructionView(LoginRequiredMixin, StaffOnlyMixin, TemplateView):
    template_name = 'dashboard/instruction.html'


class BaseExportView(View):
    """Базовый класс для выгрузки данных."""
    model = None
    filename = None
    header = None

    def get_queryset(self):
        q = self.request.GET.get("search")
        if q:
            return self.model.objects.filter(name__icontains=search)
        return self.model.objects.all()

    def get_data(self):
        """Возвращает данные для выгрузки."""
        queryset = self.get_queryset()
        data = []
        for obj in queryset:
            row = []
            for attr in self.header:
                if hasattr(obj, attr):
                    val = getattr(obj, attr)
                    if callable(val):
                        val = val()
                    row.append(str(val))
                else:
                    row.append('-')
            data.append(row)
        return data


class BrigadeCSVExportView(LoginRequiredMixin, StaffOnlyMixin, BaseExportView):
    """Выгрузка данных бригады в CSV."""
    model = Brigade
    filename = 'brigades.csv'
    header = ['id', 'name', 'description', 'customer', 'notes']

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'

        writer = csv.writer(response)
        writer.writerow(self.header)
        writer.writerows(self.get_data())
        return response


class BrigadeExcelExportView(LoginRequiredMixin, StaffOnlyMixin, BaseExportView):
    """Выгрузка данных бригады в Excel."""
    model = Brigade
    filename = 'brigades.xlsx'
    header = ['id', 'name', 'description', 'customer', 'notes']

    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(self.header)
        for row in self.get_data():
            sheet.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        workbook.save(response)
        return response


class EquipmentCSVExportView(LoginRequiredMixin, StaffOnlyMixin, BaseExportView):
    """Выгрузка данных оборудования в CSV."""
    model = Equipment
    filename = 'equipments.csv'
    header = ['id', 'serial', 'name', 'category', 'brigade', 'condition', 'date_release', 'date_exploitation',
              'manufacturer', 'certificate_start', 'certificate_end']

    def get_data(self):
        """Возвращает данные для выгрузки."""
        queryset = self.get_queryset()
        data = []
        for obj in queryset:
            row = [
                obj.id,
                obj.serial,
                obj.name,
                obj.category.name,
                obj.brigade.name,
                obj.get_condition_display(),
                obj.date_release,
                obj.date_exploitation,
                obj.manufacturer,
                obj.certificate_start,
                obj.certificate_end
            ]
            data.append(row)
        return data

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'

        writer = csv.writer(response)
        writer.writerow(self.header)
        writer.writerows(self.get_data())
        return response


class EquipmentExcelExportView(LoginRequiredMixin, StaffOnlyMixin, View):
    """Выгрузка данных оборудования в Excel."""
    filename = 'equipments.xlsx'
    header = ['id', 'serial', 'name', 'category', 'brigade', 'condition', 'date_release', 'date_exploitation',
              'manufacturer', 'certificate_start', 'certificate_end']

    def get(self, request, *args, **kwargs):

        search_request = request.GET.get("search")
        if search_request:
            equipment_by_name = Equipment.objects.filter(name__icontains=search_request)
            equipment_by_serial = Equipment.objects.filter(serial__icontains=search_request)
            queryset = (equipment_by_name | equipment_by_serial)
        else:
            queryset = Equipment.objects.all()

        data = []
        for obj in queryset:
            row = [
                obj.id,
                obj.serial,
                obj.name,
                obj.category.name,
                obj.brigade.name,
                obj.get_condition_display(),
                obj.date_release,
                obj.date_exploitation,
                obj.manufacturer,
                obj.certificate_start,
                obj.certificate_end
            ]
            data.append(row)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(self.header)
        for row in data:
            sheet.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        workbook.save(response)
        return response


class BrigadeActivityExcelView(LoginRequiredMixin, StaffOnlyMixin, View):
    def get(self, request, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = int(request.GET.get('year', datetime.now().year))

        if not 1 <= month <= 12:
            month = datetime.now().month
        if not 1900 <= year <= datetime.now().year + 10:
            year = datetime.now().year

        brigades = Brigade.objects.all().order_by('name')
        days_in_month = calendar.monthrange(year, month)[1]
        days = get_days_in_month(month, year)

        # Create a new workbook and add a worksheet.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Brigade Activities - {calendar.month_name[month]} {year}"

        # Add headers
        headers = ['Бригада'] + [f'{day}' for day in days] + ['Итого']
        worksheet.append(headers)

        # Define color mapping for work types
        color_mapping = {
            '(ЗБС) Бурение': 'FFFF0000',  # Red
            '(ЗБС) бурение горизонта': 'FF00FF00',  # Green
            '(ЗБС) шаблонирование + ГИС': 'FF0000FF',  # Blue
            '(ЗБС) спуск хвостовика': 'FFFFFF00',  # Yellow
            '(ВНС) бурение кондуктора': 'FF00FFFF',  # Cyan
            '(ВНС) бурение ЭК': 'FFFF00FF',  # Magenta
            '(ВНС) бурение горизонта': 'FFC0C0C0',  # Light Gray
            '(ВНС) спуск хвостовика': 'FF808080',  # Dark Gray
            'Переезд': 'FFA52A2A',  # Brown
            'Простой': 'FFFFC0CB',  # Pink
            'Авария': 'FF800000',  # Maroon
        }

        # Populate data
        for brigade in brigades:
            row_data = [brigade.name]
            total_ba = 0
            for day in days:
                ba = BrigadeActivity.objects.filter(brigade=brigade, date__day=day, date__month=month,
                                                    date__year=year).last()
                try:
                    activity_str = ba.work_object.short_name
                except:
                    activity_str = ""
                row_data.append(activity_str)
                if ba:
                    total_ba += 1  # Increment the total count for the brigade

            row_data.append(total_ba)  # Add the total to the row
            worksheet.append(row_data)

        # Add formatting (optional)
        # Example: Bold the header row
        for cell in worksheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=brigade_activities_{month}_{year}.xlsx'

        # Save the workbook to the response
        workbook.save(response)

        return response


class WorkerActivityExcelView(LoginRequiredMixin, StaffOnlyMixin, View):
    def get(self, request, *args, **kwargs):
        month = int(request.GET.get('month', datetime.now().month))
        year = int(request.GET.get('year', datetime.now().year))

        if not 1 <= month <= 12:
            month = datetime.now().month
        if not 1900 <= year <= datetime.now().year + 10:
            year = datetime.now().year

        brigades = Brigade.objects.all().order_by('name')
        days_in_month = calendar.monthrange(year, month)[1]
        days = get_days_in_month(month, year)

        users = User.objects.filter(is_staff=True).order_by('last_name', 'first_name')

        # Create a new workbook and add a worksheet.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Brigade Activities - {calendar.month_name[month]} {year}"

        # Add headers
        headers = ['ФИО'] + ['Должность'] + ['Бригада'] + [f'{day}' for day in days] + ['Итого']
        worksheet.append(headers)

        # Populate data
        for user in users:
            row_data = [user.get_full_name(), user.profile.position, user.profile.brigade.__str__()]
            total_ba = 0
            for day in days:
                wa = WorkerActivity.objects.filter(user=user, date__day=day, date__month=month, date__year=year).last()
                try:
                    activity_str = wa.work_type
                except:
                    activity_str = ""
                row_data.append(activity_str)
                if wa:
                    total_ba += 1  # Increment the total count for the brigade

            row_data.append(total_ba)  # Add the total to the row
            worksheet.append(row_data)

        # Add formatting (optional)
        # Example: Bold the header row
        for cell in worksheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=brigade_activities_{month}_{year}.xlsx'

        # Save the workbook to the response
        workbook.save(response)

        return response


class VehicleMovementExcelExportView(LoginRequiredMixin, StaffOnlyMixin, View):
    def get(self, request, *args, **kwargs):
        form = VehicleMovementFilterForm(request.GET)
        queryset = VehicleMovement.objects.all()

        if form.is_valid():
            month = form.cleaned_data.get('month')
            year = form.cleaned_data.get('year')
            brigade_from = form.cleaned_data.get('brigade_from')
            brigade_to = form.cleaned_data.get('brigade_to')
            driver = form.cleaned_data.get('driver')
            vehicle = form.cleaned_data.get('vehicle')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            if month:
                queryset = queryset.filter(date__month=month)
            if year:
                queryset = queryset.filter(date__year=year)
            if brigade_from:
                queryset = queryset.filter(brigade_from=brigade_from)
            if brigade_to:
                queryset = queryset.filter(brigade_to=brigade_to)
            if driver:
                queryset = queryset.filter(driver=driver)
            if vehicle:
                queryset = queryset.filter(vehicle=vehicle)
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__lte=end_date)

        # Create a workbook and add a worksheet.
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicle Movements"

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        def apply_header_style(cell):
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        def apply_cell_style(cell):
            cell.border = border

        # Write headers
        headers = [
            'Дата', 'ФИО Водителя', 'Автомобиль', 'Из бригады', 'В бригаду', 'Оборудование'
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            apply_header_style(cell)
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20

        # Write data
        for row_num, movement in enumerate(queryset, start=2):
            worksheet.cell(row=row_num, column=1, value=movement.date.strftime('%d.%m.%Y'))
            if movement.driver:
                worksheet.cell(row=row_num, column=2, value=f'{movement.driver.last_name} {movement.driver.first_name}')
            else:
                worksheet.cell(row=row_num, column=2, value='---')
            worksheet.cell(row=row_num, column=3, value=str(movement.vehicle) if movement.vehicle else '---')
            worksheet.cell(row=row_num, column=4, value=movement.brigade_from.name if movement.brigade_from else '---')
            worksheet.cell(row=row_num, column=5, value=movement.brigade_to.name if movement.brigade_to else '---')

            equipment_list = '\n'.join(
                [f'{e.equipment.name} ({e.quantity} шт.)' for e in movement.vehiclemovementequipment_set.all()])
            worksheet.cell(row=row_num, column=6, value=equipment_list)
            for col in range(1, 7):
                apply_cell_style(worksheet.cell(row=row_num, column=col))

        # Add analytics sheet
        analytics_worksheet = workbook.create_sheet(title="Analytics")

        # Add Total Movements Count
        analytics_worksheet.cell(row=1, column=1, value='Total Movements')
        apply_header_style(analytics_worksheet.cell(row=1, column=1))
        analytics_worksheet.cell(row=1, column=2, value=queryset.count())
        apply_cell_style(analytics_worksheet.cell(row=1, column=2))

        # Movements by Month
        analytics_worksheet.cell(row=3, column=1, value='Movements by Month')
        apply_header_style(analytics_worksheet.cell(row=3, column=1))
        analytics_worksheet.cell(row=3, column=2, value='Month')
        apply_header_style(analytics_worksheet.cell(row=3, column=2))
        analytics_worksheet.cell(row=3, column=3, value='Count')
        apply_header_style(analytics_worksheet.cell(row=3, column=3))

        movements_by_month = queryset.annotate(month_year=TruncMonth('date')).values('month_year').annotate(
            count=Count('id')).order_by('month_year')

        for row_num, item in enumerate(movements_by_month, start=4):
            analytics_worksheet.cell(row=row_num, column=2, value=item['month_year'].strftime('%m.%Y'))
            analytics_worksheet.cell(row=row_num, column=3, value=item['count'])
            for col in range(2, 4):
                apply_cell_style(analytics_worksheet.cell(row=row_num, column=col))

        # Movements by Brigade From
        analytics_worksheet.cell(row=len(movements_by_month) + 6, column=1, value='Movements by Brigade From')
        apply_header_style(analytics_worksheet.cell(row=len(movements_by_month) + 6, column=1))
        analytics_worksheet.cell(row=len(movements_by_month) + 6, column=2, value='Brigade')
        apply_header_style(analytics_worksheet.cell(row=len(movements_by_month) + 6, column=2))
        analytics_worksheet.cell(row=len(movements_by_month) + 6, column=3, value='Count')
        apply_header_style(analytics_worksheet.cell(row=len(movements_by_month) + 6, column=3))

        movements_by_brigade_from = queryset.values('brigade_from__name').annotate(count=Count('id')).order_by(
            '-count')

        for row_num, item in enumerate(movements_by_brigade_from, start=len(movements_by_month) + 7):
            analytics_worksheet.cell(row=row_num, column=2, value=item['brigade_from__name'] or '---')
            analytics_worksheet.cell(row=row_num, column=3, value=item['count'])
            for col in range(2, 4):
                apply_cell_style(analytics_worksheet.cell(row=row_num, column=col))

        # Movements by Brigade To
        analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=1,
                                 value='Movements by Brigade To')
        apply_header_style(
            analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=1))
        analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=2,
                                 value='Brigade')
        apply_header_style(
            analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=2))
        analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=3,
                                 value='Count')
        apply_header_style(
            analytics_worksheet.cell(row=len(movements_by_month) + len(movements_by_brigade_from) + 9, column=3))

        movements_by_brigade_to = queryset.values('brigade_to__name').annotate(count=Count('id')).order_by(
            '-count')

        for row_num, item in enumerate(movements_by_brigade_to,
                                       start=len(movements_by_month) + len(movements_by_brigade_from) + 10):
            analytics_worksheet.cell(row=row_num, column=2, value=item['brigade_to__name'] or '---')
            analytics_worksheet.cell(row=row_num, column=3, value=item['count'])
            for col in range(2, 4):
                apply_cell_style(analytics_worksheet.cell(row=row_num, column=col))

        # Movements by Driver
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=1,
            value='Movements by Driver')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=1))
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=2,
            value='Driver')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=2))
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=3,
            value='Count')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + 12, column=3))

        movements_by_driver = queryset.values('driver__last_name', 'driver__first_name').annotate(
            count=Count('id')).order_by('-count')

        for row_num, item in enumerate(movements_by_driver,
                                       start=len(movements_by_month) + len(movements_by_brigade_from) + len(
                                               movements_by_brigade_to) + 13):
            analytics_worksheet.cell(row=row_num, column=2,
                                     value=f"{item['driver__last_name']} {item['driver__first_name']}")
            analytics_worksheet.cell(row=row_num, column=3, value=item['count'])
            for col in range(2, 4):
                apply_cell_style(analytics_worksheet.cell(row=row_num, column=col))

        # Movements by Vehicle
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=1, value='Movements by Vehicle')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=1))
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=2, value='Vehicle')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=2))
        analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=3, value='Count')
        apply_header_style(analytics_worksheet.cell(
            row=len(movements_by_month) + len(movements_by_brigade_from) + len(movements_by_brigade_to) + len(
                movements_by_driver) + 15, column=3))

        movements_by_vehicle = queryset.values('vehicle__brand', 'vehicle__model', 'vehicle__number').annotate(
            count=Count('id')).order_by('-count')

        for row_num, item in enumerate(movements_by_vehicle,
                                       start=len(movements_by_month) + len(movements_by_brigade_from) + len(
                                               movements_by_brigade_to) + len(movements_by_driver) + 16):
            analytics_worksheet.cell(row=row_num, column=2,
                                     value=f"{item['vehicle__brand']} {item['vehicle__model']} {item['vehicle__number']}")
            analytics_worksheet.cell(row=row_num, column=3, value=item['count'])
            for col in range(2, 4):
                apply_cell_style(analytics_worksheet.cell(row=row_num, column=col))

        workbook.save(output)
        # Construct the response object
        filename = 'vehicle_movements.xlsx'
        response = HttpResponse(output.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


from openpyxl.utils import get_column_letter
